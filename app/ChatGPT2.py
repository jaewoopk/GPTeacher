import openai
import time
import openpyxl
# API 키

OPENAI_API_KEY = "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"
openai.api_key = OPENAI_API_KEY

model = "gpt-3.5-turbo"

messages = [{"role": "system", "content" :"Make 10 fill-in-the-blank TOEFL vocabulary problems with 4 choices. provide answer."}]
# 엑셀 파일

fpath = '/Users/KJH/Documents/GitHub/Capstone_6/excel/Sentence2.xlsx'
wb = openpyxl.load_workbook(fpath)
# 워크시트
ws = wb['문장모음']
k = 101
for n in range(100):

    k = k + 25
    for i in range(5):
        if n % 4 == 0:
            content = "Make 10 fill-in-the-blank TOEFL vocabulary problems with 4 choices. provide answer."
        elif n % 4 == 1:
            content = "Make 10 fill-in-the-blank TOEIC vocabulary problems with 4 choices. provide answer."
        elif n % 4 == 2:
            content = "Make 10 fill-in-the-blank IELTS vocabulary problems with 4 choices. provide answer."
        else:
            content = "Make 10 fill-in-the-blank TEPS vocabulary problems with 4 choices. provide answer."
        messages.append({"role": "user", "content": content})

        print(messages)
        completion = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=messages
        )

        chat_response = completion.choices[0].message.content
    # chat_response 는 str 변수이다.
        print(chat_response)
        chat_response = chat_response.replace('Answer:', '').replace('-', '').replace(')', ' ').replace('.','').replace('Explanation:')
        sentenceNum=0
        for M in range(1,10):

            if chat_response.split('\n')[M] != "":
                if chat_response.split('\n')[M].find('__') != -1:
                    sentenceNum = M
        for N in range(10):
            chat_response = chat_response.replace(str(N), '')

        print(sentenceNum)
        for j in range(10):
            ws['A'+str(k)] = chat_response.split('\n')[0+sentenceNum*j]
            if sentenceNum == 8:
                ws['F'+str(k)] = chat_response.split('\n')[6+sentenceNum*j].split(' ')[0]
            else:
                ws['F' + str(k)] = chat_response.split('\n')[5 + sentenceNum * j].split(' ')[0]
            ws['B'+str(k)] = chat_response.split('\n')[1+sentenceNum*j].split(' ')[1]
            ws['C'+str(k)] = chat_response.split('\n')[2+sentenceNum*j].split(' ')[1]
            ws['D'+str(k)] = chat_response.split('\n')[3+sentenceNum*j].split(' ')[1]
            ws['E'+str(k)] = chat_response.split('\n')[4+sentenceNum*j].split(' ')[1]

            k = k + 1
        #print(chat_response.split('.')[4])
        messages.append({"role": "assistant", "content": chat_response})
        wb.save('/Users/KJH/Documents/GitHub/Capstone_6/excel/Sentence2.xlsx')


