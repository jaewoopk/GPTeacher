import openai
import time
import openpyxl
import copy
# API 키

OPENAI_API_KEY = "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"
openai.api_key = OPENAI_API_KEY

model = "gpt-3.5-turbo"

message1 = [{"role": "user", "content" :"Make 5 fill-in-the-blank TOEFL vocabulary problems with 4 choices. provide answer.provide Explanation."}]
message2 = [{"role": "user", "content" :"Make 5 fill-in-the-blank TOEIC vocabulary problems with 4 choices. provide answer.provide Explanation."}]
message3 = [{"role": "user", "content" :"Make 5 fill-in-the-blank IELTS vocabulary problems with 4 choices. provide answer.provide Explanation."}]
message4 = [{"role": "user", "content" :"Make 5 fill-in-the-blank GRE vocabulary problems with 4 choices. provide answer.provide Explanation."}]
message5 = [{"role": "user", "content": "Make 5 English fill-in-the-blank grammar problems related to preposition with 4 choices. provide answer.provide Explanation."}]
message6 = [{"role": "user", "content": "Make 5 English fill-in-the-blank grammar problems related to noun with 4 choices. provide answer.provide Explanation."}]
message7 = [{"role": "user", "content": "Make 5 English fill-in-the-blank grammar problems related to verb with 4 choices.provide answer.provide Explanation."}]
message8 = [{"role": "user", "content": "Make 5 English fill-in-the-blank grammar problems related to Choosing Adjectives and Adverbs with 4 choices. provide answer.provide Explanation."}]
message9 = [{"role": "user", "content": "Make 5 english word meaning matching problem with 4 choices. provide answer.provide Explanation."}]
## noun 카테고리 제외
message10 = [{"role": "user", "content": "Make 5 English fill-in-the-blank grammar problems related to subject object complement with 4 choices.provide answer.provide Explanation."}]
## message 10 주어 목적어 보어 자리 관련

# 엑셀 파일
#       messages[0]['content'] = messages[0]['content'] + 'sadsadadasd' 난이도 추가 방식 If the overall difficulty is 5, it is about 1 difficulty
fpath = '/Users/KJH/Documents/GitHub/Capstone_6/excel/Sentence.xlsx'
wb = openpyxl.load_workbook(fpath)
ws = wb['문장모음']

k =2478
i = 0
for s in range(50):
    while(True):             # 문장 7 개  # (6,0,-1) 역순 (7 ) 정순
        print(i)
        if i > 6:
            i = i - 7
        if i == 0:
            message = message1
        elif i == 1:
            message = message2
        elif i == 2:
            message = message3
        elif i == 3:
            message = message4
        elif i == 4:
            message = message5
        elif i == 5:
            message = message8
        elif i == 6:
            message = message7
        for j in range(4):         # 난이도별로 한번씩
            message_input = copy.deepcopy(message)
            print(message)
            message_input[0]['content'] = message_input[0]['content'] +'If the overall difficulty is 5, it has '+str(j+1)
            print(message_input)
            time.sleep(50)
            completion = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=message_input
            )
            chat_response = completion.choices[0].message.content
            print(chat_response)
            if chat_response.split('\n')[0].find('/') != -1:
                continue
            if chat_response.split('\n')[0].find('(') != -1:
                continue
            if chat_response.split('\n')[1].count(')') >= 3:
                continue
            if chat_response.split('\n')[1].count('.') >= 3:
                continue
            if chat_response.split('\n')[0].find(':') != -1:
                continue
            answer_ExpVar = 0
            for answer_Exp in range(15):
                if chat_response.split('\n')[answer_Exp].find('Answer:') != -1:
                    if chat_response.split('\n')[answer_Exp].find('Explanation:') != -1:
                        answer_ExpVar = 1
            if answer_ExpVar == 1:
                answer_ExpVar = 0
                continue
            chat_response = chat_response.replace('Answer: ', '').replace('-', '').replace(')', '').replace('.','').replace('Explanation:','').replace('\n\n', '\n')

            for N in range(5):
                chat_response = chat_response.replace('Difficulty '+str(N+1)+':\n', '')
                chat_response = chat_response.replace(str(N+1), '')
                chat_response = chat_response.replace('problem:'+str(N+1)+':\n', '')
            GorV = 0
            if i <= 3:
                GorV = 5
            else:
                GorV = 5
            print('완료-------'+str(i+1)+'->문제분류    '+str(j+1)+'->난이도    ' + str(k) +'->문장저장개수')
            sentenceNum =7
            for M in range(GorV):
                ws['A'+str(k)] = chat_response.split('\n')[0+sentenceNum*M]
                ws['B'+str(k)] = chat_response.split('\n')[1+sentenceNum*M].split(' ')[1]
                ws['C'+str(k)] = chat_response.split('\n')[2+sentenceNum*M].split(' ')[1]
                ws['D'+str(k)] = chat_response.split('\n')[3+sentenceNum*M].split(' ')[1]
                ws['E'+str(k)] = chat_response.split('\n')[4+sentenceNum*M].split(' ')[1]
                ws['F' + str(k)] = chat_response.split('\n')[5 + sentenceNum * M].split(' ')[0]
                ws['G'+str(k)] = chat_response.split('\n')[6+sentenceNum*M]
                if i == 5:
                    ws['H' + str(k)] = 8
                else:
                    ws['H'+str(k)] = i+1   # 카테고리
                ws['I'+str(k)] = j+1   # 난이도
                k = k + 1
            wb.save('/Users/KJH/Documents/GitHub/Capstone_6/excel/Sentence.xlsx')
        i = i + 1