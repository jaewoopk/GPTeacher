import openai
import time
import openpyxl
# API 키

OPENAI_API_KEY = "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"
openai.api_key = OPENAI_API_KEY

model = "gpt-3.5-turbo"

messages = [{"role": "user", "content" :"Make 5 fill-in-the-blank TOEFL vocabulary problems with 4 choices. provide answer.provide Explanation."},
            {"role": "user", "content" :"Make 5 fill-in-the-blank TOEIC vocabulary problems with 4 choices. provide answer.provide Explanation."},
            {"role": "user", "content" :"Make 5 fill-in-the-blank IELTS vocabulary problems with 4 choices. provide answer.provide Explanation."},
            {"role": "user", "content" :"Make 5 english word meaning matching problem with 4 choices. provide answer.provide Explanation."},
            {"role": "user", "content": "Make 2 English grammar problems related to tenses with 4 choices. provide answer.provide Explanation."},
            {"role": "user", "content": "Make 2 English grammar problems related to articles with 4 choices. provide answer.provide Explanation."},
            {"role": "user", "content": "Make 2 grammar problems about which verb to put in place of the verb with 4 choices.provide answer.provide Explanation."},]
# 엑셀 파일
#       messages[0]['content'] = messages[0]['content'] + 'sadsadadasd' 난이도 추가 방식 If the overall difficulty is 5, it is about 1 difficulty
fpath = '/Users/KJH/Documents/GitHub/Capstone_6/excel/Sentence.xlsx'
wb = openpyxl.load_workbook(fpath)
ws = wb['문장모음']
k = 1
for i in range(7):             # 문장 7 개
    for j in range(4):         # 난이도별로 한번씩
        messages[i]['content'] = messages[i]['content'] +'If the overall difficulty is 5, it is about '+str(j+1) +' difficulty'
        print(messages[i])
        completion = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=messages[i]
        )
        chat_response = completion.choices[0].message.content
        chat_response = chat_response.replace('Answer: ', '').replace('-', '').replace(')', ' ').replace('.','').replace('Explanation:')
        for N in range(5):
            chat_response = chat_response.replace(str(N), '')
        GorV = 0
        if i <= 3:
            GorV = 5
        else:
            GorV = 2
        print(chat_response)
        sentenceNum =0
        for k in range(GorV):
            ws['A'+str(k)] = chat_response.split('\n')[0+sentenceNum*j]
            if sentenceNum == 8:
                ws['F'+str(k)] = chat_response.split('\n')[6+sentenceNum*j].split(' ')[0]
            else:
                ws['F' + str(k)] = chat_response.split('\n')[5 + sentenceNum * j].split(' ')[0]
            ws['B'+str(k)] = chat_response.split('\n')[1+sentenceNum*j].split(' ')[1]
            ws['C'+str(k)] = chat_response.split('\n')[2+sentenceNum*j].split(' ')[1]
            ws['D'+str(k)] = chat_response.split('\n')[3+sentenceNum*j].split(' ')[1]
            ws['E'+str(k)] = chat_response.split('\n')[4+sentenceNum*j].split(' ')[1]



for i in range(10):

    time.sleep(10)
    completion = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=messages[i]
    )

    chat_response = completion.choices[0].message.content
# chat_response 는 str 변수이다.
    chat_response = chat_response.replace('Answer:', '.').replace('-','.')
    for N in range(10):
        chat_response = chat_response.replace(str(N+1)+'.', '.')
    print(chat_response)
    for j in range(10):
        ws['A'+str(k)] = chat_response.split('.')[j*3+1]
        ws['B'+str(k)] = chat_response.split('.')[3*(j+1)].replace(' ', '').replace('\n', '')
        k = k + 1
    #print(chat_response.split('.')[4])
    messages.append({"role": "assistant", "content": chat_response})
    wb.save('/Users/KJH/Documents/GitHub/Capstone_6/excel/Sentence2.xlsx')


