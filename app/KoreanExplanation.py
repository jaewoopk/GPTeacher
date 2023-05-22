import openai
import time
import openpyxl
import copy
# API 키

OPENAI_API_KEY = "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"
openai.api_key = OPENAI_API_KEY

model = "gpt-3.5-turbo"
fpath = '/Users/KJH/Documents/GitHub/Capstone_6/excel/Sentence.xlsx'
wb = openpyxl.load_workbook(fpath)
ws = wb['sentenceALL']
ws2 = wb['해설모음']
# ws[A12]
# 필요 정보 -> 문장 , 답 , 보기

message = [{"role": "user", "content" : ""}]
choice = ['A', 'a', 'B', 'b', 'C', 'c', 'D', 'd']
pc = ['C','C','D','D','E','E','F','F']
# message[0]['content'] = '문장' + "왜 정답이 " +'정답' + " 이지? 1개의 한국어 문장으로 설명해줘"
# message[0]['content'] = '문장' + "왜 정답이 " +'정답' + " 아니지? 1개의 한국어 문장으로 설명해줘"
i = 365
while(True) :

    for j in range(0, 8, 2):
        message[0]['content'] = ""
        message[0]['content'] = str(ws['B' + str(i + 1)].value) + " 왜 이 문장 빈칸의 정답이 "
        check = 0
        present_c = pc[j]

        if str(ws['G'+str(i+1)].value) == choice[j] or str(ws['G'+str(i+1)].value) == choice[j] ==  choice[j+1]:
            answer = str(ws[present_c+str(i+1)].value)
            check = 1
        else:
            non_answer = str(ws[present_c+str(i+1)].value)
        if check == 1:
            message[0]['content'] = message[0]['content'] + answer + " 이지? 1개의 한국어 문장으로 설명해"
        else :
            message[0]['content'] = message[0]['content'] + non_answer + " 아니지? 1개의 한국어 문장으로 설명해"

        message_input = copy.deepcopy(message)
        print(message_input)
        time.sleep(40)

        completion = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=message_input
        )

        chat_response = completion.choices[0].message.content
        print(chat_response)
        if j == 0: # 보기 A
            ws['L'+str(i+1)] = chat_response
        if j == 2:
            ws['M' + str(i + 1)] = chat_response
        if j == 4:
            ws['N' + str(i + 1)] = chat_response
        if j == 6:
            ws['O'+str(i+1)] = chat_response
        wb.save('/Users/KJH/Documents/GitHub/Capstone_6/excel/Sentence.xlsx')

    i = i + 1